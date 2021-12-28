import telebot
from tqdm.contrib.telegram import tqdm, trange
import os
from dotenv import load_dotenv

from fake_useragent import UserAgent
from bs4 import BeautifulSoup
import re
import requests
import time
from random import randint

import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook

import yadisk

load_dotenv()
ua = UserAgent()
y = yadisk.YaDisk(token="AQAAAAAeeuFqAAeVCjRRWT3G8khEv1eCtEu6uY4")

TELEGRAM_TOKEN = os.getenv('TELEGRAM_TOKEN')
COOKIES = os.getenv('COOKIES')
PROXY = os.getenv('PROXY')
# CHAT_ID = os.getenv('TELEGRAM_CHAT_ID')
bot = telebot.TeleBot(TELEGRAM_TOKEN)
proxies = {
    "http":PROXY
}
b = []

@bot.message_handler(content_types=['document'])
def handle_docs(message):
    try:
        bot.reply_to(message, "Обрабатываю файл")
        file_info = bot.get_file(message.document.file_id)
        downloaded_file = bot.download_file(file_info.file_path)
        src_t = '/app/documents/' + message.document.file_name
        with open(src_t, 'wb') as new_file:
            new_file.write(downloaded_file)
        name = message.document.file_name
        bot.reply_to(message, "Начал парсинг")
        read_data(name,src_t,file_info,message)
    except Exception as e:
        bot.reply_to(message, e)


def read_data(name, src_t, file_info, message):
    bot.reply_to(message, "collect")
    workbook_name2 = f'/app/documents/{name}'  
    wb2 = load_workbook(workbook_name2)
    ws2 = wb2.active
    sheet = wb2.get_sheet_by_name('Sheet1')
    # df_init = len(pd.read_excel(f'/app/documents/{name}'))+2
    df_init = 10
    chat_id = message.chat.id
    counter = 1
    for i in trange(df_init, token=TELEGRAM_TOKEN, chat_id=chat_id):
        time.sleep(randint(9,15))
        counter+=1
        res = sheet.cell(row=counter, column=1).value
        bot.reply_to(message, f"{res}")
        collect_data(name, src_t, file_info, counter, df_init, res, message)

def collect_data(name, src_t, file_info, counter, df_init, res, message):
    res1 = res
    res = res.replace(' ', '+')
    # bot.reply_to(message, f"{res} {res1}")
    url = f'https://www.ozon.ru/search/?from_global=true&text={res}'
    # bot.reply_to(message, f"{url}")
    response = requests.get(
                url=url,
                headers={'user-agent': f'{ua.random}',
                'Cookie': f'{COOKIES}'},
                proxies = proxies,
            )
    src = response.text

    try:
        # time.sleep(4)
        soup_split = BeautifulSoup(src, 'lxml')
        script = str(soup_split.find_all("script", attrs= {"type": "application/javascript"})[1])
        splited = script.split('>')[1].split('"')[1].replace("\/",'/').split('category_was_predicted=true')[0]
        url1 = f'https://www.ozon.ru{splited}category_was_predicted=true&from_global=true&text={res}'
        response1 = requests.get(
                url=url1,
                headers={'user-agent': f'{ua.random}',
                'Cookie': f'{COOKIES}'},
                proxies = proxies,
            )
        src1 = response1.text
        soup1 = BeautifulSoup(src1, 'lxml')
        try:
            search = soup1.find_all(class_="b6r7")[0].text.replace(res1,'')
            qty = ''.join(re.findall('\d', search))
            b.append(qty)
        except:
            search = soup1.find_all(class_="b3a1")[0].text.replace(res1,'')
            qty = ''.join(re.findall('\d', search))
            b.append(qty)
    except:
        try:
            soup_usualy = BeautifulSoup(src, 'lxml')
            search = soup_usualy.find_all(class_="b6r7")[0].text.replace(res1,'')
            qty = ''.join(re.findall('\d', search))
            b.append(qty)
        except:
            qty = 0
            b.append(qty)
    bot.reply_to(message, f"{qty}")
    if counter == df_init+1:
        data = {'Data': b}
        saver(name, src_t, file_info, data, message)
        b.clear()

def saver(name, src_t, file_info, data, message):
    workbook_name1 = f'/app/documents/{name}' 
    df = pd.DataFrame(data)
    writer = pd.ExcelWriter(workbook_name1, engine='xlsxwriter')
    df.to_excel(writer,startcol = 1, sheet_name='Sheet1',index=False, header=False)
    writer.save()
    y.upload(src_t, f'{file_info.file_path}')
    d_link = y.get_download_link(file_info.file_path)
    bot.reply_to(message, f"Всё готово, вот ссылка на скачивание {d_link}")
bot.infinity_polling()